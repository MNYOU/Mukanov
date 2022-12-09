import csv
import decimal
import re
import sys

import numpy as np
import pdfkit
from jinja2 import Environment, FileSystemLoader
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.utils import get_column_letter
from prettytable import prettytable


class DataSet:
    """Класс для работы с входными данными"""

    def csv_parse(self, file_name):
        """Считывает ваканси с файла, формирует их список

        Args:
            file_name (str): Название файла для чтения

        Returns:
            list: Список вакансий
        """
        with open(file_name, 'r', encoding="utf-8-sig") as csvfile:
            reader = csv.DictReader(csvfile)
            vacancies = []
            for vacancy in reader:
                if len(vacancy) == len(reader.fieldnames) and not any(
                        value is None or value == '' for value in vacancy.values()):
                    vacancy_with_correct_value = self.get_correct_vacancy(vacancy)
                    vacancies.append(Vacancy(vacancy_with_correct_value, False))
        return vacancies

    def csv_parse_for_table(self, file_name):
        """Считывает ваканси с файла, формирует их список для создания таблицы

        Args:
            file_name (str): Название файла для чтения

        Returns:
            list: Список вакансий и названий полей
        """
        with open(file_name, 'r', encoding="utf-8-sig") as csvfile:
            reader = csv.DictReader(csvfile)
            vacancies = []
            for vacancy in reader:
                if len(vacancy) == len(reader.fieldnames) and not any(
                        value is None or value == '' for value in vacancy.values()):
                    vacancy_with_correct_value = self.get_correct_vacancy(vacancy)
                    vacancies.append(Vacancy(vacancy_with_correct_value, True))
            if reader.fieldnames is None or len(vacancies) == 0:
                if reader.fieldnames is None:
                    print('Пустой файл')
                else:
                    print('Нет данных')
                sys.exit()

        return vacancies, reader.fieldnames

    def get_correct_vacancy(self, vacancy):
        """Удаляет лишние символы из вакансии

        Args:
            vacancy (dict): Словарь, значения которого нужно очистить от лишних символов

        Returns:
            dict: Вакансия с корректными значениями
        """

        def get_correct_string(s):
            """Удаляет лишние пробелы и html теги из строки

            Args:
                s (str): Строка

            Returns:
                str: Очищенная строка
            """
            s = re.sub(r'<[^>]*>', '', s)
            result = []
            for item in s.split('\n'):
                result.append(' '.join(item.split()))
            return '\n'.join(result)

        return {key: get_correct_string(vacancy[key]) for key in vacancy}


def filter_vacancies(vacancies, key, value):
    """Фильтрует вакансии по ключу и значению

    Args:
        vacancies (list): Список вакансий
        key (str): Ключ, по которому происходит фильтрация
        value (str or int): Значение для фильтрации

    Returns:
        list: Отфильтрованные вакансии
    """
    return list(filter(lambda v: v.is_suitable(key, value), vacancies))


def sort_vacancies(vacancies, key, reverse):
    """Сортирует вакансии по ключу

    Args:
        vacancies (list): Список вакансий
        key (str): Ключ, по которому происходит сортировка
        reverse (bool): Показывает, нужен ли обратный порядок сортировки

    Returns:
        bool: Отсортированные вакансии
    """
    vacancies.sort(key=lambda v: v.get_value_for_sort(key), reverse=reverse)


class Vacancy:
    """Класс для представления вакансии

    Attributes:
        name (str): Название
        description (str): Описание
        salary (Salary): Зарплата
        area_name (str): Город, в котором была размещена вакансия
        published_at (str): Время, в которое была создана вакансия
        key_skills (str): Ключевые навыки
        experience_id (str): Опыт, который должны иметь кандидаты
        premium (str): Является ли вакансия премиумной
        employer_name (str): Название кампании, разместившей вакансию
    """

    def __init__(self, data, is_for_table):
        """Инициализирует объект Salary

        Args:
            data (dict): Словарь с данными
            is_for_table (bool): Показывает, используется ли класс для создания таблицы
        """
        if data is None:
            return
        self.name = data['name']
        self.salary = Salary({key: data[key] for key in data if 'salary' in key}, is_for_table)
        self.area_name = data['area_name']
        self.published_at = data['published_at']
        if is_for_table:
            self.description = data['description']
            self.key_skills = data['key_skills'].split('\n')
            self.experience_id = data['experience_id']
            self.premium = data['premium']
            self.employer_name = data['employer_name']

    value_to_rus = {'premium': {'false': 'Нет', 'true': 'Да'},
                    'experience_id': {'noexperience': 'Нет опыта', 'between1and3': 'От 1 года до 3 лет',
                                      'between3and6': 'От 3 до 6 лет', 'morethan6': 'Более 6 лет', }}
    naming_to_en = {'Название': 'name', 'Описание': 'descriprion', 'Навыки': 'key_skills',
                    'Опыт работы': 'experience_id',
                    'Премиум-вакансия': 'premium', 'Компания': 'employer_name',
                    'Оклад': 'salary', 'Идентификатор валюты оклада': 'salary_currency',
                    'Название региона': 'area_name', 'Дата публикации вакансии': 'published_at', }

    def is_suitable(self, key, value, year_only=False):
        """Проверяет, удовлетворяет ли вакансия заданому условию

        Args:
            key (str): Ключ, по которому происходит сравнение
            value (int or str): Значение, которому должно соответствовать значение вакансии
            year_only (bool): Показывает, нужно ли сравнивать published_at только по году

        Returns:
            bool: Удовлетворяет ли вакансия заданому условию"""
        key = self.naming_to_en[key]
        if key == 'name':
            return value in self.name
        if key == 'published_at':
            if year_only:
                return value == self.published_at.split('-')[0]
            return value == self.get_date(self.published_at)
        elif key == 'salary' or key == 'salary_currency':
            return self.salary.is_suitable(key, value)
        self_value = self.__getattribute__(key)
        if key in self.value_to_rus:
            self_value = self.value_to_rus[key][self_value.lower()]
        return self_value == value

    def get_value_for_sort(self, key):
        """Подбирает значение для сортировки

        Args:
            key (str): Ключ, по которому происходит выбор значения

        Returns:
            str, int: Значение для сравнения"""
        key = self.naming_to_en[key]
        if key == 'salary':
            return self.salary.get_salary_in_rub()
        elif key == 'key_skills':
            return len(self.key_skills)
        elif key == 'experience_id':
            d = {'noexperience': 0, 'between1and3': 1, 'between3and6': 2, 'morethan6': 3}
            return d[self.experience_id.lower()]
        else:
            return self.__getattribute__(key)

    def get_formatted_value(self):
        """Генерирует значение в специальном формате при помощи словаря - value_to_rus

        Returns:
            list: Список атрибутов класса в форматном выводе"""
        f_value = [self.name,
                   self.description,
                   '\n'.join(self.key_skills),
                   self.value_to_rus['experience_id'][self.experience_id.lower()],
                   self.value_to_rus['premium'][self.premium.lower()],
                   self.employer_name,
                   self.salary.get_formatted_value(),
                   self.area_name,
                   self.get_date(self.published_at), ]

        for i, value in enumerate(f_value):
            if len(value) > 100:
                f_value[i] = value[:100] + '...'
        return f_value

    def get_date(self, s):
        """Извлекает дату из строки, содержащей дату и время

        Args:
            s (str): Дата и время в виде строки
        Returns:
            str: Дата в правильном формате"""
        time = s.split('T')[0].split('-')
        time.reverse()
        return '.'.join(time)


class Salary:
    """Класс для представления зарплаты

    Attributes:
        salary_from (int): Нижняя граница вилки оклада
        salary_to (int): Верхняя граница вилки оклада
        salary_currency (str): Индификатор валюты оклада оклада
        salary_gross (str): Показывает, учитывается ли налог
    """
    currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                       "UAH": 1.64, "USD": 60.66, "UZS": 0.0055, }

    currency_to_ru = {'azn': 'Манаты', 'byr': 'Белорусские рубли', 'eur': 'Евро', 'gel': 'Грузинский лари',
                      'kgs': 'Киргизский сом', 'kzt': 'Тенге', 'rur': 'Рубли', 'uah': 'Гривны', 'usd': 'Доллары',
                      'uzs': 'Узбекский сум'}

    gross_to_ru = {'false': 'С вычетом налогов', 'true': 'Без вычета налогов'}

    def __init__(self, dic, is_for_table):
        """Инициализирует объект Salary

        Args:
            dic (dict): Словарь с данными
            is_for_table (bool): Показывает, используется ли класс для создания таблицы
        """
        self.salary_from = dic['salary_from']
        self.salary_to = dic['salary_to']
        self.salary_currency = dic['salary_currency']
        if is_for_table:
            self.salary_gross = dic['salary_gross']

    def is_suitable(self, key, value):
        """Проверяет, удовлетворяет ли зарплата заданому условию

        Args:
            key (str): Ключ, по которому происходит сравнение
            value (int or str): Значение, которому должна соответствовать зарплата

        Returns:
            bool: Удовлетворяет ли зарплата заданому условию
        """
        if key == 'salary':
            return int(self.salary_from) <= int(value) <= int(
                self.salary_to)
        else:
            return self.currency_to_ru[self.salary_currency.lower()] == value

    def get_salary_in_rub(self):
        """Вычисляет среднюю зарплату из вилки и переводит в рубли, при помощи словаря - currency_to_rub

        Returns:
            int: Зарплата в рублях"""
        rate = self.currency_to_rub[self.salary_currency]
        return (int(self.salary_from.split('.')[0]) * rate + int(self.salary_to.split('.')[0]) * rate) / 2

    def get_formatted_value(self):
        """Генерирует значение в специальном формате при помощи словарей - currency_to_rub, gross_to_ru

        Returns:
            str: атрибуты класса в форматном выводе
        """
        salary = f'{self.get_formatted_salary(self.salary_from)} - {self.get_formatted_salary(self.salary_to)}'
        currency = f'({self.currency_to_ru[self.salary_currency.lower()]})'
        gross = f'({self.gross_to_ru[self.salary_gross.lower()]})'
        return f'{salary} {currency} {gross}'

    def get_formatted_salary(self, salary):
        """Генерирует зарплату в специальном формате

        Args:
            salary (int): Зарплата

        Returns:

            str: зарплата в специальном формате"""
        n = decimal.Decimal(salary)
        res = '{0:,}'.format(n).replace(',', ' ')
        return res.split('.')[0]


class InputConnect:
    """Класс для организации данных и создания таблицы

    Attributes:
        need_filter (bool): Показывает, нужна ли фильтрация
        key_filter (str): Ключ, по которому будет производиться фильтрация
        value_filter (str): Значение, по которому будет производиться фильтрация
        need_sort (bool): Показывает, нужна ли сортировка
        key_sort (str): Ключ, по которому будет производиться сортировка
        start (int): Номер первой вакансии в таблице
        end (int): Номер последней вакансии в таблице
        all_fields (list): Все поля, которые могут быть выведены в таблицу
    """

    def __init__(self):
        """Инициализирует объект InputConnect"""
        self.need_filter = False
        self.key_filter, self.value_filter = '', ''
        self.key_sort = ''
        self.need_sort, self.sort_reverse = False, False
        self.start = 1
        self.end = None
        self.all_fields = ['Название', 'Описание', 'Навыки', 'Опыт работы', 'Премиум-вакансия', 'Компания', 'Оклад',
                           'Название региона', 'Дата публикации вакансии']
        self.naming = self.all_fields.copy()

    def check_and_parse_input(self, param_filter, param_sort, param_reverse, numbers, naming):
        """Проверяет на коректность и парсит входные данные для таблицы, добавляет атрибуты в класс

        Args:
            param_filter (str): Ключ и значение для фильтрации, разделенные ', '
            param_sort (str): Ключ для сортировки
            param_reverse (str): Показывает, нужна ли обратная сортировка
            numbers (list): Список с начальным и конечным номерами вакансий
            naming (list): Список колонок, которые будут в таблице"""
        self.pars_filter(param_filter)
        self.pars_sort(param_sort, param_reverse)

        if len(naming) != 0 and naming[0] != '':
            self.naming = naming

        if len(numbers) == 2:
            self.start, self.end = map(int, numbers)
        elif len(numbers) == 1:
            self.start = int(numbers[0])

    def pars_filter(self, param_filter):
        """Парсит параметр фильтрации таблицы

        Args:
            param_filter (str): Ключ и значение для фильтрации, разделенные ', '
        """
        if param_filter == '':
            self.need_filter = False
            return
        if ':' not in param_filter:
            print('Формат ввода некорректен')
            sys.exit()
        key_filter, value_filter = param_filter.split(': ')
        if key_filter not in self.all_fields and key_filter != 'Идентификатор валюты оклада':
            print('Параметр поиска некорректен')
            sys.exit()
        self.need_filter = True
        self.key_filter, self.value_filter = key_filter, value_filter

    def pars_sort(self, param_sort, param_reverse):
        """Парсит параметры сортировки таблицы

        Args:
            param_sort (str): Ключ для сортировки
            param_reverse (str): Показывает, нужна ли обратная сортировка
        """
        if param_sort == '':
            return
        if param_sort not in self.all_fields:
            print('Параметр сортировки некорректен')
            sys.exit()
        if param_reverse != '' and param_reverse != 'Да' and param_reverse != 'Нет':
            print('Порядок сортировки задан некорректно')
            sys.exit()
        self.need_sort = True
        self.key_sort = param_sort
        if param_reverse == 'Да':
            self.sort_reverse = True
        else:
            self.sort_reverse = False

    def print_table(self, fields):
        """Печатает таблицу в консоль

        Args:
            fields (list): Данные, которые будут в таблице
        """
        table = self.config_table()
        for i, v in enumerate(fields):
            table.add_row([i + 1] + v.get_formatted_value())
        if self.end is None:
            self.end = len(fields) + 1
        print(table.get_string(start=self.start - 1, end=self.end - 1, fields=['№'] + self.naming))

    def config_table(self):
        """Конфигурирует таблицу

        Returns:
            Возвращает сконфигурированную таблицу
        """
        table = prettytable.PrettyTable()
        table.hrules = prettytable.ALL
        table.field_names = ['№'] + self.all_fields
        table.align = 'l'
        table.max_width = 20
        return table


class DataStats:
    """Класс для формирования статистики

    Attributes:
        prof_name (str) : Название профессии для формирования статистики
        salary_years (dict): Словарь с зарплатами по годам
        count_years (dict): Словарь с количеством вакансий по годам
        salary_prof (dict): Словарь с зарплатами по годам для выбранной профессии
        count_prof (dict): Словарь с количеством вакансий по годам для выбранной профессии
        areas_with_salrs (dict): Словарь с зарплатами по городам
        areas_with_shares (dict): Словарь с долей вакансий по городам
    """

    def __init__(self):
        """Инициализирует объект DataStats"""
        self.prof_name = ""
        self.salary_years = {}
        self.count_years = {}
        self.salary_prof = {}
        self.count_prof = {}
        self.areas_with_salrs = {}
        self.areas_with_shares = {}

    def calculate_stats(self, vacancies, prof_name):
        """Формирует статистики по годам и по городам

        Args:
            vacancies (list): Вакансии для формирования статистики
            prof_name (str): Название Профессии для формирования статистики
        """
        self.prof_name = prof_name
        for i in range(2007, 2023):
            fields = self.filter_vacancies(vacancies, 'Дата публикации вакансии', str(i), True)
            if len(fields) != 0:
                self.set_value_dicts(self.salary_years, self.count_years, i, fields)
                fields = self.filter_vacancies(fields, 'Название', self.prof_name)
                self.set_value_dicts(self.salary_prof, self.count_prof, i, fields)
        self.calculate_stats_areas(vacancies)

    def set_value_dicts(self, dic_salary, dic_count, key, vacancies):
        """Вычисляет статистику по годам и заполняет ей словари

        Args:
            dic_salary (dict): Словарь с зарплатами по годам
            dic_count (dict): Словарь с количеством вакансий по годам
            key (str or int): Ключ, по которому будут заполняться словари
            vacancies (list): Список вакансий, по которым будет вычисляться статистика
        """
        dic_salary[key] = self.get_avg_salary(vacancies)
        dic_count[key] = len(vacancies)

    def calculate_stats_areas(self, vacancies):
        """Вычисляет статистику по городам и заполняет ей атрибуты класса

        Args:
            vacancies (list): Список вакансий, по которым будет вычисляться статистика
        """
        dic_areas = {}
        for vacancy in vacancies:
            if vacancy.area_name not in dic_areas:
                dic_areas[vacancy.area_name] = 0
            dic_areas[vacancy.area_name] += 1
        areas = [area for area in dic_areas if dic_areas[area] / len(vacancies) >= 0.01]
        for area in areas:
            fields = self.filter_vacancies(vacancies, 'Название региона', area)
            self.set_value_dicts(self.areas_with_salrs, self.areas_with_shares, area, fields)
            self.areas_with_shares = {area: float(format(dic_areas[area] / len(vacancies), '.4f')) for area
                                      in self.areas_with_shares}

        self.areas_with_salrs = self.get_sorted_dic(self.areas_with_salrs, lambda item: item[1])
        self.areas_with_shares = self.get_sorted_dic(self.areas_with_shares, lambda item: item[1])

    def filter_vacancies(self, vacancies, key, value, year_only=False):
        """Фильтрует вакансии

        Args:
            vacancies (list): вакансии, требующие сортировки
            key (str): Ключ, по которому будет производиться фильтрация
            value (int or str): Значение, по которому будет производиться фильтрация
            year_only (bool): Показывает, нужно ли сравнить поле вакансии 'published_at' только по году

        Returns:
            list: Списко отфильтрованных вакансий
        """
        return list(filter(lambda v: v.is_suitable(key, value, year_only), vacancies))

    def get_avg_salary(self, vacancies):
        """Возвращает среднюю зарплату

        Args:
            vacancies (list): Вакансии, содержащие атрибут зарплаты

        Returns:
            int: Средняя зарплата в рублях
        """
        if len(vacancies) == 0:
            return 0
        return int(sum([v.salary.get_salary_in_rub() for v in vacancies]) / len(vacancies))

    def get_sorted_dic(self, dic, handler):
        """Сортирует словарь

        Args:
            dic (dict): Словарь, требующий сортировки
            handler (function): Обработчик, по которому сортируется словарь

        Returns:
            dict: Отсортированный словарь
        """
        sorted_tuples = sorted(dic.items(), key=handler, reverse=True)
        return {pair[0]: pair[1] for i, pair in enumerate(sorted_tuples) if i < 10}

    def get_all(self):
        """Возвращает всю статистику, сформированную классом

        Returns:
            dict: Словарь со статистикой
        """
        return {'salary_years': self.salary_years, 'count_years': self.count_years, 'salary_prof': self.salary_prof,
                'count_prof': self.count_prof, 'areas_with_salrs': self.areas_with_salrs,
                'areas_with_shares': self.areas_with_shares, 'prof_name': self.prof_name,
                'areas': list(self.areas_with_salrs.keys())}

    def print(self):
        """Печатает все атрибуты класса в консоль"""
        print('Динамика уровня зарплат по годам:', self.salary_years)
        print('Динамика количества вакансий по годам:', self.count_years)
        print('Динамика уровня зарплат по годам для выбранной профессии:', self.salary_prof)
        print('Динамика количества вакансий по годам для выбранной профессии:', self.count_prof)
        print('Уровень зарплат по городам (в порядке убывания):', self.areas_with_salrs)
        print('Доля вакансий по городам (в порядке убывания):', self.areas_with_shares)


class Report:
    """Класс для формирования отчетов

    Attributes:
        book (Workbook): excel книга
        sheet_cities (Worksheet): лист excel книги со статистикой по годам
        sheet_years (Worksheet): лист excel книги со статистикой по городам
        data (dict): словарь со статистикой
    """

    def __init__(self, dicts):
        """Инициализирует объект Report

        Args:
            dicts (dict): Словарь, содержащий словари со статистикой
        """
        self.book = None
        self.sheet_cities = None
        self.sheet_years = None
        self.data = dicts

    def generate_excel(self):
        """Генерирует excel файл и заполняет его данными"""
        book = Workbook()
        self.book = book
        book.remove(book.active)
        self.sheet_years = self.fill_data_years(book.create_sheet('Статистика по годам'))
        self.sheet_cities = self.fill_data_cities(book.create_sheet('Статистика по городам'))
        self.stylize_book(book)
        book.save('report.xlsx')
        return book

    def fill_data_years(self, sheet):
        """Заполняет excel лист статистикой по годам

        Args:
            sheet (Worksheet): Excel лист

        Returns:
            (Worksheet): заполненный данными excel лист
        """
        sheet_years = [
            ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.data["prof_name"]}', 'Количество вакансий',
             f'Количество вакансий - {self.data["prof_name"]}']]
        sheet.append(sheet_years[0])
        years = list(self.data['salary_years'].keys())
        for i in range(years[0], years[-1] + 1):
            row = [i, self.data['salary_years'][i], self.data['salary_prof'][i], self.data['count_years'][i],
                   self.data['count_prof'][i]]
            sheet.append(row)
            sheet_years.append(row)
        return sheet_years

    def fill_data_cities(self, sheet):
        """Заполняет excel лист статистикой по городам

        Args:
            sheet (Worksheet): Excel лист

        Returns:
            (Worksheet): заполненный данными excel лист
        """
        sheet_cities = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        sheet.append(sheet_cities[0])
        values = []
        for key, value in self.data['areas_with_salrs'].items():
            values.append([key, value, ''])
        for i, (key, value) in enumerate(self.data['areas_with_shares'].items()):
            values[i] += [key, value]
        for row in values:
            sheet.append(row)
            sheet_cities.append(row)
        return sheet_cities

    def stylize_book(self, book):
        """Стилизует excel книгу

        Args:
            book (Workbook): Excel книга
        """
        bold_style = self.get_base_style('bold_style')
        bold_style.font = Font(bold=True)
        book.add_named_style(bold_style)
        book.add_named_style(self.get_base_style('normal_style'))
        percent_style = self.get_base_style('percent_style')
        percent_style.number_format = BUILTIN_FORMATS[10]
        book.add_named_style(percent_style)

        self.stylize_worksheet(book['Статистика по годам'])
        self.stylize_worksheet(book['Статистика по городам'], is_percent_value=True, column_index=4)
        self.make_width_correct(book)

    def get_base_style(self, name):
        """Создает именованный стиль и делает его базовую настройку

        Args:
            name (str): Название стиля

        Returns:
            (NamedStyle): Именованный стиль
        """
        style = NamedStyle(name=name)
        side = Side(style='thin', color='000000')
        style.border = Border(top=side, left=side, right=side, bottom=side)
        return style

    def stylize_worksheet(self, sheet, is_percent_value=False, column_index=-1):
        """Стилизует excel лист

        Args:
            sheet (Worksheet): Excel лист
            is_percent_value (bool): Показывает, содержит ли лист столбы с '%'
            column_index (int): Индекс стоблца, содержащего '%'
        """
        for i, row in enumerate(sheet):
            for j, cell in enumerate(row):
                if i == 0:
                    cell.style = 'bold_style'
                elif is_percent_value and j == column_index:
                    cell.style = 'percent_style'
                else:
                    cell.style = 'normal_style'

    def make_width_correct(self, workbook):
        """Задает корректрую ширину всем колонкам книги

        Args:
            workbook (Workbook): Excel книга
        """
        for worksheet in workbook.worksheets:
            column_widths = []
            for row in worksheet:
                for i, cell in enumerate(row):
                    len_value = len(cell.value) if isinstance(cell.value, str) else len(str(cell.value))
                    if len(column_widths) > i:
                        if len_value > column_widths[i]:
                            column_widths[i] = len_value
                    else:
                        column_widths += [len_value]

            for i, column_width in enumerate(column_widths, 1):
                worksheet.column_dimensions[get_column_letter(i)].width = column_width + 2

    def generate_image(self):
        """Создает диаграмму и сохраняет её в png файл"""
        plt.figure()
        plt.rcParams['font.size'] = '8'

        self.get_base_chart('Уровень зарплат по годам', 1, 'средняя з/п', f'з/п {self.data["prof_name"]}',
                            self.data['salary_years'], self.data['salary_prof'])
        self.get_base_chart('Количество вакансий по годам', 2, 'Количество вакансий',
                            f'Количество вакансий {self.data["prof_name"]}', self.data['count_years'],
                            self.data['count_prof'])
        self.get_barhchart()
        self.get_piechart()

        plt.tight_layout()
        plt.savefig('graph.png')

    def get_base_chart(self, title, position, label1, label2, data1, data2):
        """Рисует базовую столбчатую диаграмму с двумя графиками

        Args:
            title (str): Название диаграммы
            position (int): позиция диаграммы в общей фигуре
            label1 (str): Подписи к первому графику
            label2 (str): Подписи к второму графику
            data1 (dict): Данные для первого графика
            data2 (dict): Данные для второго графика
        """
        labels = list(data1.keys())
        labels_indexes = np.array(labels)
        data1 = list(data1.values())
        data2 = list(data2.values())
        width = 0.4
        plt.subplot(2, 2, position)
        plt.title(title)
        plt.bar(labels_indexes - width / 2, data1, label=label1, width=width)
        plt.bar(labels_indexes + width / 2, data2, label=label2, width=width)
        plt.grid(axis='y')
        plt.xticks(rotation=90)
        plt.legend()

    def get_barhchart(self):
        """Рисует перевернутую столбчатую диаграмму"""
        plt.subplot(2, 2, 3)
        plt.title('Уровень зарплат по городам')
        areas = list(map(lambda x: x.replace(' ', ' \n').replace('-', '-\n'), self.data['areas']))
        areas.reverse()
        salaries = list(self.data['areas_with_salrs'].values()).copy()
        salaries.reverse()
        plt.barh(areas, salaries)
        plt.tick_params(axis='y', which='major', labelsize=6)
        plt.grid(axis='x')

    def get_piechart(self):
        """Рисует квуговую диаграмму"""
        plt.subplot(2, 2, 4)
        areas = self.data['areas'].copy()
        plt.title('Доля вакансий по городам')
        percents = list(self.data['areas_with_shares'].values())
        areas.insert(0, 'Другие')
        percents.insert(0, 1 - sum(percents))
        plt.pie(percents, labels=areas, textprops={'fontsize': 6})

    def generate_pdf(self):
        """Генерирует pdf файл, используя excel и png отчеты"""
        pdf_template = self.create_template()
        # config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltox\bin\wkhtmltopdf.exe')
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)

    def create_template(self):
        """Создает html шаблон pdf файла"""
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        headings_years, data_years = self.get_formatted_data('Статистика по годам')
        headings_cities, data_cities = self.get_formatted_data('Статистика по городам', True)
        pdf_template = template.render(
            {'headings_years': headings_years, 'data_years': data_years,
             'headings_cities': headings_cities, 'data_cities': data_cities})
        return pdf_template

    def get_formatted_data(self, sheet_name, need_formatting=False):
        """Возвращает данные для html шаблона в специальном формате

        Args:
            sheet_name (str): Название листа, из которого будут браться данные
            need_formatting (bool): Показывает, нужно ли добавлять к ячейкам '%'

        Returns:
            (list): Список названий столбцов и данные
        """
        data = []
        is_heading = True
        for row in self.book[sheet_name]:
            if is_heading:
                naming = list(map(lambda x: x.value, row))
                is_heading = False
            else:
                row_values = list(map(lambda x: x.value, row))
                if need_formatting:
                    row_values[-1] = format(row_values[-1], '.2%')
                data.append(row_values)
        return naming, data


def start_data_to_table():
    """Запускает сценарий получения вакансий в виде таблице"""
    file_name = input('Введите название файла: ')
    input_connect = InputConnect()
    input_connect.check_and_parse_input(
        input('Введите параметр фильтрации: '),
        input('Введите параметр сортировки: '),
        input('Обратный порядок сортировки (Да / Нет): '),
        input('Введите диапазон вывода: ').split(),
        input('Введите требуемые столбцы: ').split(', '))
    vacancies, source_naming = DataSet().csv_parse_for_table(file_name)
    if input_connect.need_filter:
        vacancies = filter_vacancies(vacancies, input_connect.key_filter, input_connect.value_filter)
        if len(vacancies) == 0:
            print('Ничего не найдено')
            sys.exit()
    if input_connect.need_sort:
        sort_vacancies(vacancies, input_connect.key_sort, input_connect.sort_reverse)

    input_connect.print_table(vacancies)


def start_data_to_stats():
    """Запускает сценарий получения статистики в виде pdf файла"""
    file_name = input('Введите название файла: ')
    prof_name = input('Введите название профессии: ')
    data_set = DataSet()
    vacancies = data_set.csv_parse(file_name)

    data_stats = DataStats()
    data_stats.calculate_stats(vacancies, prof_name)
    data_stats.print()

    report = Report(data_stats.get_all())
    report.generate_excel()
    report.generate_image()
    report.generate_pdf()


s = input('Какие данные вы хотели бы видеть(Вакансии/Статистика)?: ')
if s == 'Вакансии':
    start_data_to_table()
elif s == 'Статистика':
    start_data_to_stats()
