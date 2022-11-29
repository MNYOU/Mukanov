import re
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS


class DataSet:
    def csv_parse(self, file_name):
        with open(file_name, 'r', encoding="utf-8-sig") as csvfile:
            reader = csv.DictReader(csvfile)
            vacancies = []
            for vacancy in reader:
                if len(vacancy) == len(reader.fieldnames) and not any(
                        value is None or value == '' for value in vacancy.values()):
                    vacancy_with_correct_value = self.get_correct_vacancy(vacancy)
                    vacancies.append(Vacancy(vacancy_with_correct_value))
        return vacancies

    def get_correct_vacancy(self, vacancy):
        def get_correct_string(s):
            s = re.sub(r'<[^>]*>', '', s)
            result = []
            for item in s.split('\n'):
                result.append(' '.join(item.split()))
            return '\n'.join(result)

        return {k: get_correct_string(vacancy[k]) for k in vacancy}


class Vacancy:
    def __init__(self, data):
        if data is None:
            return
        self.name = data['name']
        self.salary = Salary({key: data[key] for key in data if 'salary' in key})
        self.area_name = data['area_name']
        self.published_at = data['published_at']

    value_to_rus = {'premium': {'false': 'Нет', 'true': 'Да'},
                    'experience_id': {'noexperience': 'Нет опыта', 'between1and3': 'От 1 года до 3 лет',
                                      'between3and6': 'От 3 до 6 лет', 'morethan6': 'Более 6 лет', }}
    naming_to_en = {'Название': 'name', 'Описание': 'descriprion', 'Навыки': 'key_skills',
                    'Опыт работы': 'experience_id',
                    'Премиум-вакансия': 'premium', 'Компания': 'employer_name',
                    'Оклад': 'salary', 'Идентификатор валюты оклада': 'salary_currency',
                    'Название региона': 'area_name', 'Дата публикации вакансии': 'published_at', }

    def is_suitable(self, key, value):
        key = self.naming_to_en[key]
        if key == 'name':
            return value in self.name
        if key == 'published_at':
            return value == self.get_time(self.published_at)
        elif key == 'salary' or key == 'salary_currency':
            return self.salary.is_suitable(key, value)
        self_value = self.__getattribute__(key)
        if key in self.value_to_rus:
            self_value = self.value_to_rus[key][self_value.lower()]
        return self_value == value

    def get_value_for_sort(self, key):
        key = self.naming_to_en[key]
        if key == 'salary':
            return self.salary.get_value_for_compare()
        else:
            return self.__getattribute__(key)

    def get_time(self, s):
        return s.split('-')[0]


class Salary:
    currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                       "UAH": 1.64, "USD": 60.66, "UZS": 0.0055, }

    currency_to_ru = {'azn': 'Манаты', 'byr': 'Белорусские рубли', 'eur': 'Евро', 'gel': 'Грузинский лари',
                      'kgs': 'Киргизский сом', 'kzt': 'Тенге', 'rur': 'Рубли', 'uah': 'Гривны', 'usd': 'Доллары',
                      'uzs': 'Узбекский сум'}

    gross_to_ru = {'false': 'С вычетом налогов', 'true': 'Без вычета налогов'}

    def __init__(self, dic):
        self.salary_from = dic['salary_from']
        self.salary_to = dic['salary_to']
        self.salary_currency = dic['salary_currency']

    def is_suitable(self, key, value):
        if key == 'salary':
            return int(self.salary_from) <= int(value) <= int(
                self.salary_to)
        else:
            return self.currency_to_ru[self.salary_currency.lower()] == value

    def salary_in_rub(self):
        rate = self.currency_to_rub[self.salary_currency]
        return int(self.salary_from.split('.')[0]) * rate, int(self.salary_to.split('.')[0]) * rate

    def get_value_for_compare(self):
        return sum(self.salary_in_rub()) / 2


class DataStats:
    def __init__(self):
        self.salary_years = {}
        self.count_years = {}
        self.salary_prof = {}
        self.count_prof = {}
        self.areas = []
        self.areas_with_salrs = {}
        self.areas_with_shares = {}

    def calculate_stats(self, vacancies, name):
        for i in range(2007, 2023):
            fields = self.filter_vacancies(vacancies, 'Дата публикации вакансии', str(i))
            if len(fields) != 0:
                self.set_value_dicts(self.salary_years, self.count_years, i, fields)
                fields = self.filter_vacancies(fields, 'Название', name)
                self.set_value_dicts(self.salary_prof, self.count_prof, i, fields)
        self.calculate_stats_areas(vacancies)

    def set_value_dicts(self, dic_salary, dic_count, key, fields):
        dic_salary[key] = self.get_avg_salary(fields)
        dic_count[key] = len(fields)

    def calculate_stats_areas(self, vacancies):
        dic_areas = {}
        for vacancy in vacancies:
            if vacancy.area_name not in dic_areas:
                dic_areas[vacancy.area_name] = 0
            dic_areas[vacancy.area_name] += 1
        areas = [area for area in dic_areas if dic_areas[area] / len(vacancies) >= 0.01]
        # areas = areas[:10] города ведь не отсортированы
        for area in areas:
            fields = self.filter_vacancies(vacancies, 'Название региона', area)
            self.set_value_dicts(self.areas_with_salrs, self.areas_with_shares, area, fields)
            self.areas_with_shares = {area: float(format(dic_areas[area] / len(vacancies), '.4f')) for area
                                      in self.areas_with_shares}

        self.areas_with_salrs = self.get_sorted_dic(self.areas_with_salrs, lambda item: item[1])
        self.areas_with_shares = self.get_sorted_dic(self.areas_with_shares, lambda item: item[1])

    def filter_vacancies(self, vacancies, key, value):
        return list(filter(lambda v: v.is_suitable(key, value), vacancies))

    def get_avg_salary(self, vacancies):
        if len(vacancies) == 0:
            return 0
        return int(sum([v.salary.get_value_for_compare() for v in vacancies]) / len(vacancies))

    def get_sorted_dic(self, dic, handler):
        sorted_tuples = sorted(dic.items(), key=handler, reverse=True)
        return {pair[0]: pair[1] for i, pair in enumerate(sorted_tuples) if i < 10}

    def get_all(self):
        return {'salary_years': self.salary_years, 'count_years': self.count_years, 'salary_prof': self.salary_prof,
                'count_prof': self.count_prof, 'areas_with_salrs': self.areas_with_salrs,
                'areas_with_shares': self.areas_with_shares, }

    def print(self):
        print('Динамика уровня зарплат по годам:', self.salary_years)
        print('Динамика количества вакансий по годам:', self.count_years)
        print('Динамика уровня зарплат по годам для выбранной профессии:', self.salary_prof)
        print('Динамика количества вакансий по годам для выбранной профессии:', self.count_prof)
        print('Уровень зарплат по городам (в порядке убывания):', self.areas_with_salrs)
        print('Доля вакансий по городам (в порядке убывания):', self.areas_with_shares)


class Report:
    def __init__(self):
        self.data = {}

    def generate_excel(self, dicts):
        self.data = dicts
        book = openpyxl.Workbook()
        book.remove(book.active)
        self.fill_data_years(book.create_sheet('Статистика по годам'))
        self.fill_data_cities(book.create_sheet('Статистика по городам'))
        self.stylize_book(book)
        book.save('report.xlsx')

    def fill_data_years(self, sheet):
        sheet.append(['Год', 'Средняя зарплата', f'Средняя зарплата - {prof_name}', 'Количество вакансий',
                      f'Количество вакансий - {prof_name}'])
        years = list(self.data['salary_years'].keys())
        for i in range(years[0], years[-1] + 1):
            row = [i, self.data['salary_years'][i], self.data['salary_prof'][i], self.data['count_years'][i],
                   self.data['count_prof'][i]]
            sheet.append(row)

    def fill_data_cities(self, sheet):
        sheet.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        values = []
        for key, value in self.data['areas_with_salrs'].items():
            values.append([key, value, ''])
        for i, (key, value) in enumerate(self.data['areas_with_shares'].items()):
            values[i] += [key, value]
        for row in values:
            sheet.append(row)

    def stylize_book(self, book):
        bold_style = self.get_base_style('bold_style')
        bold_style.font = Font(bold=True)
        book.add_named_style(bold_style)
        book.add_named_style(self.get_base_style('normal_style'))
        percent_style = self.get_base_style('percent_style')
        percent_style.number_format = BUILTIN_FORMATS[10]
        book.add_named_style(percent_style)

        self.stylize_worksheet(book['Статистика по годам'])
        self.stylize_worksheet(book['Статистика по городам'], is_percent_value=True, column_index=4)
        self.make_width_correct(book)

    def get_base_style(self, name):
        style = NamedStyle(name=name)
        side = Side(style='thin', color='000000')
        style.border = Border(top=side, left=side, right=side, bottom=side)
        return style

    def stylize_worksheet(self, sheet, is_percent_value=False, column_index=-1):
        for i, row in enumerate(sheet):
            for j, cell in enumerate(row):
                if i == 0:
                    cell.style = 'bold_style'
                elif is_percent_value and j == column_index:
                    cell.style = 'percent_style'
                else:
                    cell.style = 'normal_style'

    def make_width_correct(self, workbook):
        for worksheet in workbook.worksheets:
            column_widths = []
            for row in worksheet:
                for i, cell in enumerate(row):
                    len_value = len(cell.value) if isinstance(cell.value, str) else len(str(cell.value))
                    if len(column_widths) > i:
                        if len_value > column_widths[i]:
                            column_widths[i] = len_value
                    else:
                        column_widths += [len_value]

            for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
                worksheet.column_dimensions[get_column_letter(i)].width = column_width + 2


file_name = input('Введите название файла: ')
prof_name = input('Введите название профессии: ')
data_set = DataSet()
vacancies = data_set.csv_parse(file_name)

data_stats = DataStats()
data_stats.calculate_stats(vacancies, prof_name)
data_stats.print()

report = Report()
report.generate_excel(data_stats.get_all())
